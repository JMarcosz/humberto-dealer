"""Servicio de importación y exportación Excel con openpyxl."""
import logging
import os
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# Columnas esperadas en el Excel de importación (orden no importa, se busca por header)
COLUMNAS_REQUERIDAS = {"modelo", "anio", "vin", "color", "precio", "combustible", "transmision"}
COLUMNAS_OPCIONALES = {"kilometraje", "descripcion"}

# Mapeo del formato Humberto: nombre en Excel → (marca, modelo, categoria)
BRAND_MAP = {
    'MERCEDES BENZ':             ('MERCEDES BENZ', 'Clase C',        'SEDAN'),
    'RANGE ROVER':               ('LAND ROVER',    'Range Rover',    'SUV'),
    'RANGE ROVER FREELANDER':    ('LAND ROVER',    'Freelander',     'SUV'),
    'BMW':                       ('BMW',           'Serie 3',        'SEDAN'),
    'BMW X5':                    ('BMW',           'X5',             'SUV'),
    'JEEP G CHEROKE':            ('JEEP',          'Grand Cherokee', 'SUV'),
    'JEEP GRAND CHEROKEE':       ('JEEP',          'Grand Cherokee', 'SUV'),
    'JEEP COMPAS':               ('JEEP',          'Compass',        'SUV'),
    'JEEP COMPASS':              ('JEEP',          'Compass',        'SUV'),
    'JEEP PATRIO':               ('JEEP',          'Patriot',        'SUV'),
    'FORD SCAPE':                ('FORD',          'Escape',         'SUV'),
    'FORD ESCAPE':               ('FORD',          'Escape',         'SUV'),
    'FORD EXPLORER':             ('FORD',          'Explorer',       'SUV'),
    'FORD FIESTA':               ('FORD',          'Fiesta',         'SEDAN'),
    'FORD VENZA':                ('FORD',          'Venza',          'SUV'),
    'FORD F-150':                ('FORD',          'F-150',          'PICKUP'),
    'CAMIONETA FORD F-150':      ('FORD',          'F-150',          'PICKUP'),
    'FORD FOCUS':                ('FORD',          'Focus',          'SEDAN'),
    'HYUNDAI SANTAFE':           ('HYUNDAI',       'Santa Fe',       'SUV'),
    'HYUNDAI SANTA FE':          ('HYUNDAI',       'Santa Fe',       'SUV'),
    'HYUNDAI TUZON':             ('HYUNDAI',       'Tucson',         'SUV'),
    'HYUNDAI TUCSON':            ('HYUNDAI',       'Tucson',         'SUV'),
    'HYUNDAI NEW TUCSON':        ('HYUNDAI',       'Tucson',         'SUV'),
    'HYUNDAI':                   ('HYUNDAI',       'Elantra',        'SEDAN'),
    'HYUNDAI LF':                ('HYUNDAI',       'Elantra',        'SEDAN'),
    'HYUNDAI I20':               ('HYUNDAI',       'i20',            'SEDAN'),
    'MAZDA CX9':                 ('MAZDA',         'CX-9',           'SUV'),
    'MAZDA CX7':                 ('MAZDA',         'CX-7',           'SUV'),
    'MAZDA CX-7':                ('MAZDA',         'CX-7',           'SUV'),
    'MAZDA CX8':                 ('MAZDA',         'CX-8',           'SUV'),
    'MAZDA':                     ('MAZDA',         'CX-7',           'SUV'),
    'DOGE G CARAVAN':            ('DODGE',         'Grand Caravan',  'VAN'),
    'DOGE  G CARAVAN':           ('DODGE',         'Grand Caravan',  'VAN'),
    'DOGE CARAVAN':              ('DODGE',         'Caravan',        'VAN'),
    'DOGE  CARAVAN':             ('DODGE',         'Caravan',        'VAN'),
    'DODGE DURANGO':             ('DODGE',         'Durango',        'SUV'),
    'TOYOTA FOR RONER':          ('TOYOTA',        '4Runner',        'SUV'),
    'TOYOTA VENZA':              ('TOYOTA',        'Venza',          'SUV'),
    'TOYOTA TACOMA':             ('TOYOTA',        'Tacoma',         'PICKUP'),
    'TOYOTA HIGHLANDER':         ('TOYOTA',        'Highlander',     'SUV'),
    'SUZUKI VITARA':             ('SUZUKI',        'Vitara',         'SUV'),
    'SUZUKI GRAND VITARA':       ('SUZUKI',        'Grand Vitara',   'SUV'),
    'CHEVROLET EQUINOX':         ('CHEVROLET',     'Equinox',        'SUV'),
    'CHEVROLET EQUINOX LIMITED': ('CHEVROLET',     'Equinox',        'SUV'),
    'CHEVROLET EQUINOX PREMIER': ('CHEVROLET',     'Equinox',        'SUV'),
    'CHEVROLET MALIBU':          ('CHEVROLET',     'Malibu',         'SEDAN'),
    'CHEVROLET SPARK':           ('CHEVROLET',     'Spark',          'SEDAN'),
    'CHEVROLET TRAX':            ('CHEVROLET',     'Trax',           'SUV'),
    'GMC':                       ('GMC',           'Acadia',         'SUV'),
    'MITSUBISHI ENDEAVOR':       ('MITSUBISHI',    'Endeavor',       'SUV'),
    'MITSUBISHI MONTERO':        ('MITSUBISHI',    'Montero',        'SUV'),
    'MITSUBISHI CARA DE GATO':   ('MITSUBISHI',    'Eclipse',        'COUPE'),
    'NISSAN SENTRA':             ('NISSAN',        'Sentra',         'SEDAN'),
    'NISSAN PATHFINDER':         ('NISSAN',        'Pathfinder',     'SUV'),
    'NISSAN NAVARA':             ('NISSAN',        'Navara',         'PICKUP'),
    'NISSAN ROGUE':              ('NISSAN',        'Rogue',          'SUV'),
    'NISSAN NOTE VERSION S':     ('NISSAN',        'Note',           'SEDAN'),
    'NISSAN NOTE VERSIÓN S':     ('NISSAN',        'Note',           'SEDAN'),
    'HONDA PILOT':               ('HONDA',         'Pilot',          'SUV'),
    'HONDA CR-V':                ('HONDA',         'CR-V',           'SUV'),
    'HONDA FIT':                 ('HONDA',         'Fit',            'SEDAN'),
    'HONDA CIVIC':               ('HONDA',         'Civic',          'SEDAN'),
    'KIA SOUL':                  ('KIA',           'Soul',           'SEDAN'),
    'KIA SPORTAGE':              ('KIA',           'Sportage',       'SUV'),
    'KIA K5':                    ('KIA',           'K5',             'SEDAN'),
    'ACURA TL':                  ('ACURA',         'TL',             'SEDAN'),
    'VOLKSWAGEN TOUAREG':        ('VOLKSWAGEN',    'Touareg',        'SUV'),
}
SKIP_MARCAS = {'LÍNEA VACÍA', 'LINEA VACIA', 'CAMION HYUNDAI FURGON', 'CAMOIN HYUNDAI FURGOM'}


class ExcelService:

    # ----------------------------------------------------------
    # Leer Excel → lista de dicts + lista de errores por línea
    # ----------------------------------------------------------
    @staticmethod
    def leer_vehiculos(ruta: str) -> tuple[list[dict], list[str]]:
        filas:   list[dict] = []
        errores: list[str]  = []

        try:
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.active

            # Detectar fila de headers (puede estar en fila 1 o más abajo)
            header_row = 1
            col_idx: dict = {}
            for row_num in range(1, 6):
                row_vals = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=row_num, max_row=row_num))]
                col_idx = {h: i for i, h in enumerate(row_vals) if h}
                if COLUMNAS_REQUERIDAS <= set(col_idx.keys()):
                    header_row = row_num
                    break
                # Detectar formato Humberto (tiene 'marca' pero no 'modelo')
                if 'marca' in col_idx and 'modelo' not in col_idx:
                    header_row = row_num
                    wb.close()
                    return ExcelService._leer_formato_humberto(ruta, header_row)
            else:
                faltantes = COLUMNAS_REQUERIDAS - set(col_idx.keys())
                if faltantes:
                    errores.append(f"Columnas faltantes en el archivo: {', '.join(sorted(faltantes))}")
                    wb.close()
                    return filas, errores

            # Normalizar columnas con tilde → sin tilde para aceptar ambas formas
            _norm = {'año': 'anio', 'transmisión': 'transmision', 'descripción': 'descripcion'}
            for src, dst in _norm.items():
                if src in col_idx and dst not in col_idx:
                    col_idx[dst] = col_idx.pop(src)

            for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                if not any(row):
                    continue
                try:
                    fila = ExcelService._parsear_fila(row, col_idx, row_num)
                    filas.append(fila)
                except ValueError as ve:
                    errores.append(str(ve))

            wb.close()
        except Exception as exc:
            errores.append(f"Error al leer el archivo: {exc}")

        return filas, errores

    @staticmethod
    def _leer_formato_humberto(ruta: str, header_row: int) -> tuple[list[dict], list[str]]:
        filas:   list[dict] = []
        errores: list[str]  = []
        vin_counter = 1

        try:
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.active

            headers = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=header_row, max_row=header_row))]
            col_idx = {h: i for i, h in enumerate(headers) if h}

            for row_num, row in enumerate(ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1):
                if not any(row):
                    continue

                def get(campo):
                    idx = col_idx.get(campo)
                    return row[idx] if idx is not None and idx < len(row) else None

                marca_raw = str(get('marca') or '').strip().upper()
                if not marca_raw or marca_raw in SKIP_MARCAS:
                    continue

                # Buscar mapping en BRAND_MAP con fallback a coincidencia parcial
                mapping = BRAND_MAP.get(marca_raw)
                if not mapping:
                    for key, val in BRAND_MAP.items():
                        if marca_raw.startswith(key) or key in marca_raw:
                            mapping = val
                            break
                if not mapping:
                    errores.append(f"Fila {row_num}: sin mapeo para '{marca_raw}' — omitida")
                    continue

                marca_nombre, modelo_nombre, categoria = mapping

                try:
                    precio = float(get('precio') or 0)
                    if precio <= 0:
                        continue
                except (TypeError, ValueError):
                    errores.append(f"Fila {row_num}: precio inválido — omitida")
                    continue

                try:
                    anio_val = get('a\xf1o') or get('ano') or get('anio') or get('año')
                    anio = int(anio_val or 0)
                    if not (1990 <= anio <= datetime.utcnow().year + 1):
                        anio = 2000
                except (TypeError, ValueError):
                    anio = 2000

                color = str(get('color') or 'OTRO').strip().upper() or 'OTRO'
                vin   = f'HAI{vin_counter:014d}'
                vin_counter += 1

                filas.append({
                    'marca_nombre':  marca_nombre,
                    'modelo':        modelo_nombre,
                    'modelo_categoria': categoria,
                    'anio':          anio,
                    'vin':           vin,
                    'color':         color,
                    'precio':        precio,
                    'combustible':   'GASOLINA',
                    'transmision':   'AUTOMATICA',
                    'kilometraje':   0,
                    'descripcion':   None,
                    'importado_excel': True,
                })

            wb.close()
        except Exception as exc:
            errores.append(f"Error al leer el archivo: {exc}")

        return filas, errores

    @staticmethod
    def _parsear_fila(row: tuple, col_idx: dict, row_num: int) -> dict:
        def get(campo: str) -> Any:
            idx = col_idx.get(campo)
            return row[idx] if idx is not None and idx < len(row) else None

        errores_fila = []

        modelo      = str(get("modelo") or "").strip().upper()
        vin         = str(get("vin") or "").strip().upper()
        color       = str(get("color") or "").strip().upper()
        combustible = str(get("combustible") or "").strip().upper()
        transmision = str(get("transmision") or "").strip().upper()

        try:
            anio = int(get("anio") or 0)
            if not (1990 <= anio <= datetime.utcnow().year + 1):
                errores_fila.append(f"Fila {row_num}: año inválido ({anio})")
        except (TypeError, ValueError):
            errores_fila.append(f"Fila {row_num}: año no numérico")
            anio = 0

        try:
            precio = float(get("precio") or 0)
            if precio <= 0:
                errores_fila.append(f"Fila {row_num}: precio debe ser positivo")
        except (TypeError, ValueError):
            errores_fila.append(f"Fila {row_num}: precio no numérico")
            precio = 0.0

        if not modelo:
            errores_fila.append(f"Fila {row_num}: modelo vacío")
        if len(vin) != 17:
            errores_fila.append(f"Fila {row_num}: VIN inválido ({vin})")

        if errores_fila:
            raise ValueError(" | ".join(errores_fila))

        return {
            "modelo":      modelo,
            "anio":        anio,
            "vin":         vin,
            "color":       color,
            "precio":      precio,
            "combustible": combustible,
            "transmision": transmision,
            "kilometraje": int(get("kilometraje") or 0),
            "descripcion": str(get("descripcion") or "").strip() or None,
        }

    # ----------------------------------------------------------
    # Exportar lista de Vehiculo → archivo .xlsx
    # ----------------------------------------------------------
    @staticmethod
    def exportar_vehiculos(vehiculos: list) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario"

        headers = [
            "ID", "Marca", "Modelo", "Año", "VIN", "Color",
            "Precio (USD)", "Km", "Combustible", "Transmisión",
            "Estado", "Publicado en",
        ]
        header_fill = PatternFill("solid", fgColor="1A1A2E")
        header_font = Font(color="FFFFFF", bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center")

        for row_num, v in enumerate(vehiculos, start=2):
            marca  = v.modelo.marca.nombre if v.modelo and v.modelo.marca else ""
            modelo = v.modelo.nombre if v.modelo else ""
            ws.append([
                v.id, marca, modelo, v.anio, v.vin, v.color,
                float(v.precio), v.kilometraje, v.combustible, v.transmision,
                v.estado, v.publicado_en.strftime("%Y-%m-%d %H:%M") if v.publicado_en else "",
            ])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True  # type: ignore

        ruta = os.path.join(os.path.dirname(__file__), "..", "..", "tmp_exports", "inventario.xlsx")
        os.makedirs(os.path.dirname(ruta), exist_ok=True)
        ruta = os.path.normpath(ruta)
        wb.save(ruta)
        log.info("Excel exportado: %d vehículos → %s", len(vehiculos), ruta)
        return ruta

    # ----------------------------------------------------------
    # Generar plantilla Excel vacía para importación
    # ----------------------------------------------------------
    @staticmethod
    def generar_plantilla() -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla"

        headers = ["modelo", "año", "vin", "color", "precio",
                   "combustible", "transmisión", "kilometraje", "descripción"]
        header_fill = PatternFill("solid", fgColor="1A1A2E")
        header_font = Font(color="FFFFFF", bold=True)

        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Fila de ejemplo
        ws.append(["Corolla", 2023, "1HGBH41JXMN109186", "BLANCO",
                   25000, "GASOLINA", "AUTOMATICA", 0, "Descripción opcional"])

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].auto_size = True  # type: ignore

        ruta = os.path.join(os.path.dirname(__file__), "..", "..", "tmp_exports", "plantilla_vehiculos.xlsx")
        os.makedirs(os.path.dirname(ruta), exist_ok=True)
        ruta = os.path.normpath(ruta)
        wb.save(ruta)
        return ruta
